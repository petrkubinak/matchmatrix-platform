#!/usr/bin/env python3
# -*- coding: utf-8 -*-

r"""
MATCHMATRIX – IMPORT HISTORICKÉHO KORPUSU DO DATABÁZE

CO:
Samostatný importér nekanonických historických dokumentů a příloh.

K ČEMU:
- načte celý archiv bez požadavku na jednotnou strukturu dokumentů,
- zachová celý vytěžený obsah,
- uloží základní metadata, původní cesty a obsahové duplicity,
- vytvoří vyhledatelné sekce,
- neovlivní kanonické dokumenty MM-DOC / MM-STD / MM-REF,
- podporuje VALIDATE_ONLY, DRY_RUN a APPLY.

KDE:
tools/documentation/25_1_A_25_IMPORT_HISTORY_CORPUS_TO_DB_V1.py

JAK:
Validace a vytvoření manifestu:
    py -3.14 .\tools\documentation\25_1_A_25_IMPORT_HISTORY_CORPUS_TO_DB_V1.py --validate-only

Bezpečný databázový test s rollbackem:
    py -3.14 .\tools\documentation\25_1_A_25_IMPORT_HISTORY_CORPUS_TO_DB_V1.py --dsn "<DSN>"

Skutečný import:
    py -3.14 .\tools\documentation\25_1_A_25_IMPORT_HISTORY_CORPUS_TO_DB_V1.py --dsn "<DSN>" --apply
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import sys
import zipfile
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable
import xml.etree.ElementTree as ET


ENGINE_VERSION = "A25_HISTORY_CORPUS_DATABASE_IMPORT_V1_1"
DEFAULT_ARCHIVE_ROOT = (
    r"\\192.168.3.119\matchmatrix\docs\99_ARCHIVE"
    r"\09_HISTORY\historie 25062026"
)
DEFAULT_ID_REGISTRY = (
    Path(__file__).resolve().parents[2]
    / "config"
    / "documentation"
    / "history_corpus_id_registry.json"
)

SUPPORTED_TEXT_EXTENSIONS = {".md", ".txt", ".sql", ".csv", ".json", ".yaml", ".yml", ".log"}
SUPPORTED_DOCUMENT_EXTENSIONS = {".docx", ".pdf", ".xlsx"}
BINARY_ATTACHMENT_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp"}
ALL_SUPPORTED_EXTENSIONS = (
    SUPPORTED_TEXT_EXTENSIONS
    | SUPPORTED_DOCUMENT_EXTENSIONS
    | BINARY_ATTACHMENT_EXTENSIONS
)

SENSITIVE_EXCLUDED_RELATIVE_PATHS = {
    "KeePass.pdf",
    "komunikace s chatGPT/02_2026/n\u00e1zev v dockeru plus heslo.png",
}
DOCUMENT_ID_RE = re.compile(r"^MM-HIS-(\d{4})$")
DATE_TOKEN_RE = re.compile(r"(?<!\d)(20\d{2})[-_]?(\d{2})[-_]?(\d{2})(?!\d)")
MARKDOWN_HEADING_RE = re.compile(r"^(#{1,6})\s+(.+?)\s*$")
OOXML_XLSX_MARKER = "xl/workbook.xml"
OOXML_DOCX_MARKER = "word/document.xml"


@dataclass
class ExtractedFile:
    absolute_path: Path
    relative_path: str
    source_extension: str
    detected_format: str
    title: str
    content_markdown: str
    extraction_status: str
    extraction_warnings: list[str]
    source_sha256: str
    content_sha256: str
    size_bytes: int
    modified_at: str
    document_date: str | None
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass
class CorpusDocument:
    document_id: str
    canonical: ExtractedFile
    variants: list[ExtractedFile]
    sections: list[dict[str, Any]]

    @property
    def metadata(self) -> dict[str, Any]:
        return {
            "corpus_engine": ENGINE_VERSION,
            "archive_root": str(self.canonical.absolute_path.parent),
            "canonical_relative_path": self.canonical.relative_path,
            "detected_format": self.canonical.detected_format,
            "source_extension": self.canonical.source_extension,
            "document_date": self.canonical.document_date,
            "content_sha256": self.canonical.content_sha256,
            "source_variants": [
                {
                    "relative_path": item.relative_path,
                    "source_sha256": item.source_sha256,
                    "source_extension": item.source_extension,
                    "detected_format": item.detected_format,
                    "size_bytes": item.size_bytes,
                    "modified_at": item.modified_at,
                    "extraction_status": item.extraction_status,
                    "extraction_warnings": item.extraction_warnings,
                }
                for item in self.variants
            ],
            "variant_count": len(self.variants),
            "content_duplicate": len(self.variants) > 1,
            "extraction_status": self.canonical.extraction_status,
            "extraction_warnings": self.canonical.extraction_warnings,
        }


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def normalize_newlines(text: str) -> str:
    return text.replace("\r\n", "\n").replace("\r", "\n").strip() + "\n"


def decode_text(data: bytes) -> tuple[str, str]:
    if data.startswith(b"\xef\xbb\xbf"):
        return data.decode("utf-8-sig"), "utf-8-sig"
    if data.startswith(b"\xff\xfe") or data.startswith(b"\xfe\xff"):
        return data.decode("utf-16"), "utf-16"

    for encoding in ("utf-8", "cp1250", "cp1252", "iso-8859-2"):
        try:
            return data.decode(encoding), encoding
        except UnicodeDecodeError:
            continue

    return data.decode("utf-8", errors="replace"), "utf-8-replacement"


def safe_title(text: str, fallback: str) -> str:
    for line in text.splitlines():
        cleaned = line.strip()
        if not cleaned:
            continue
        heading = MARKDOWN_HEADING_RE.match(cleaned)
        if heading:
            cleaned = heading.group(2).strip()
        cleaned = re.sub(r"\s+", " ", cleaned).strip(" |#\t")
        if cleaned:
            return cleaned[:300]
    return fallback[:300]


def detect_document_date(relative_path: str) -> str | None:
    match = DATE_TOKEN_RE.search(relative_path)
    if not match:
        return None
    try:
        return datetime(
            int(match.group(1)),
            int(match.group(2)),
            int(match.group(3)),
        ).date().isoformat()
    except ValueError:
        return None


def is_zip_with_marker(data: bytes, marker: str) -> bool:
    try:
        with zipfile.ZipFile(BytesIO(data)) as archive:
            return marker in archive.namelist()
    except zipfile.BadZipFile:
        return False


def markdown_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    return text.replace("|", r"\|").replace("\r", " ").replace("\n", "<br>")


def extract_xlsx(data: bytes, filename: str) -> tuple[str, dict[str, Any], list[str]]:
    warnings: list[str] = []
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Pro čtení XLSX chybí openpyxl. "
            "Nainstaluj: py -3.14 -m pip install openpyxl"
        ) from exc

    workbook = load_workbook(BytesIO(data), read_only=True, data_only=False)
    parts = ["# Tabulkov? dokument", ""]
    sheet_metadata = []

    try:
        for worksheet in workbook.worksheets:
            rows = [
                tuple(row)
                for row in worksheet.iter_rows(values_only=True)
                if any(value not in (None, "") for value in row)
            ]

            width = max((len(row) for row in rows), default=0)
            sheet_metadata.append(
                {
                    "title": worksheet.title,
                    "rows": len(rows),
                    "columns": width,
                }
            )

            parts.extend([f"## List: {worksheet.title}", ""])

            if not rows:
                parts.extend(["_Prázdný list._", ""])
                continue

            for index, row in enumerate(rows, start=1):
                values = [markdown_cell(value) for value in row]
                parts.append(f"{index:06d}\t" + "\t".join(values))
            parts.append("")
    finally:
        workbook.close()

    metadata = {"worksheets": sheet_metadata}
    return normalize_newlines("\n".join(parts)), metadata, warnings


def extract_docx(data: bytes, filename: str) -> tuple[str, dict[str, Any], list[str]]:
    warnings: list[str] = []
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}

    with zipfile.ZipFile(BytesIO(data)) as archive:
        xml_data = archive.read(OOXML_DOCX_MARKER)

    root = ET.fromstring(xml_data)
    body = root.find("w:body", namespace)
    parts = [f"# {filename}", ""]
    paragraph_count = 0
    table_count = 0

    if body is not None:
        for child in list(body):
            tag = child.tag.rsplit("}", 1)[-1]

            if tag == "p":
                texts = [node.text or "" for node in child.findall(".//w:t", namespace)]
                paragraph = "".join(texts).strip()
                if paragraph:
                    parts.append(paragraph)
                    parts.append("")
                    paragraph_count += 1

            elif tag == "tbl":
                table_count += 1
                parts.extend([f"## Tabulka {table_count}", ""])
                for row_index, row in enumerate(child.findall(".//w:tr", namespace), start=1):
                    cells = []
                    for cell in row.findall("./w:tc", namespace):
                        texts = [node.text or "" for node in cell.findall(".//w:t", namespace)]
                        cells.append(markdown_cell("".join(texts).strip()))
                    parts.append(f"{row_index:06d}\t" + "\t".join(cells))
                parts.append("")

    metadata = {
        "paragraph_count": paragraph_count,
        "table_count": table_count,
    }
    return normalize_newlines("\n".join(parts)), metadata, warnings


def extract_pdf(data: bytes, filename: str) -> tuple[str, dict[str, Any], list[str]]:
    warnings: list[str] = []
    try:
        from pypdf import PdfReader
    except ImportError:
        warnings.append("PDF_TEXT_EXTRACTION_DEPENDENCY_MISSING")
        content = normalize_newlines(
            f"# {filename}\n\n"
            "_PDF byl evidován, ale text zatím nebyl vytěžen. "
            "Pro úplnou extrakci nainstaluj balíček pypdf._\n"
        )
        return content, {"page_count": None}, warnings

    reader = PdfReader(BytesIO(data))
    parts = [f"# {filename}", ""]
    for page_number, page in enumerate(reader.pages, start=1):
        parts.extend([f"## Strana {page_number}", ""])
        text = page.extract_text() or ""
        parts.extend([text.strip() or "_Bez vytěžitelného textu._", ""])

    return (
        normalize_newlines("\n".join(parts)),
        {"page_count": len(reader.pages)},
        warnings,
    )


def extract_binary_placeholder(
    filename: str,
    extension: str,
    size_bytes: int,
) -> tuple[str, dict[str, Any], list[str]]:
    warning = "BINARY_ATTACHMENT_METADATA_ONLY"
    content = normalize_newlines(
        f"# Binární příloha: {filename}\n\n"
        f"- Formát: `{extension or 'bez přípony'}`\n"
        f"- Velikost: `{size_bytes}` B\n"
        "- Obsah nebyl v této importní vlně automaticky vytěžen.\n"
    )
    return content, {"binary_attachment": True}, [warning]


def extract_file(root: Path, path: Path) -> ExtractedFile:
    data = path.read_bytes()
    relative_path = str(path.relative_to(root)).replace("\\", "/")
    extension = path.suffix.lower()
    detected_format = extension.lstrip(".") or "binary"
    warnings: list[str] = []
    extra_metadata: dict[str, Any] = {}

    if is_zip_with_marker(data, OOXML_XLSX_MARKER):
        detected_format = "xlsx"
        content, extra_metadata, warnings = extract_xlsx(data, path.name)

    elif is_zip_with_marker(data, OOXML_DOCX_MARKER):
        detected_format = "docx"
        content, extra_metadata, warnings = extract_docx(data, path.name)

    elif extension == ".pdf":
        detected_format = "pdf"
        content, extra_metadata, warnings = extract_pdf(data, path.name)

    elif extension in SUPPORTED_TEXT_EXTENSIONS:
        text, encoding = decode_text(data)
        detected_format = extension.lstrip(".") or "text"
        content = normalize_newlines(text)
        extra_metadata = {"text_encoding": encoding}

    elif extension in BINARY_ATTACHMENT_EXTENSIONS:
        detected_format = extension.lstrip(".") or "image"
        content, extra_metadata, warnings = extract_binary_placeholder(
            path.name,
            extension,
            len(data),
        )

    else:
        content, extra_metadata, warnings = extract_binary_placeholder(
            path.name,
            extension,
            len(data),
        )
        warnings.append("UNSUPPORTED_EXTENSION_METADATA_ONLY")

    title = (
        path.stem
        if detected_format == "xlsx"
        else safe_title(content, path.stem)
    )
    content_hash = sha256_bytes(content.encode("utf-8"))
    extraction_status = "READY" if not warnings else "READY_WITH_WARNINGS"

    return ExtractedFile(
        absolute_path=path,
        relative_path=relative_path,
        source_extension=extension,
        detected_format=detected_format,
        title=title,
        content_markdown=content,
        extraction_status=extraction_status,
        extraction_warnings=warnings,
        source_sha256=sha256_bytes(data),
        content_sha256=content_hash,
        size_bytes=len(data),
        modified_at=datetime.fromtimestamp(
            path.stat().st_mtime,
            tz=timezone.utc,
        ).isoformat(),
        document_date=detect_document_date(relative_path),
        metadata=extra_metadata,
    )


def section_key(value: str, ordinal: int) -> str:
    normalized = re.sub(r"[^A-Z0-9]+", "_", value.upper()).strip("_")
    if not normalized:
        normalized = f"SECTION_{ordinal:04d}"
    return normalized[:180]


def parse_sections(content: str) -> list[dict[str, Any]]:
    lines = content.splitlines()
    heading_positions: list[tuple[int, int, str]] = []

    for index, line in enumerate(lines):
        match = MARKDOWN_HEADING_RE.match(line)
        if match:
            heading_positions.append(
                (index, len(match.group(1)), match.group(2).strip())
            )

    if not heading_positions:
        return [
            {
                "section_order": 1,
                "heading_level": 1,
                "title": "Celý obsah",
                "section_key": "FULL_CONTENT",
                "content_markdown": normalize_newlines(content),
            }
        ]

    sections: list[dict[str, Any]] = []

    if heading_positions[0][0] > 0:
        preamble = "\n".join(lines[: heading_positions[0][0]]).strip()
        if preamble:
            sections.append(
                {
                    "section_order": 1,
                    "heading_level": 1,
                    "title": "Úvodní obsah",
                    "section_key": "PREAMBLE",
                    "content_markdown": normalize_newlines(preamble),
                }
            )

    for position, (line_index, level, title) in enumerate(heading_positions):
        next_index = (
            heading_positions[position + 1][0]
            if position + 1 < len(heading_positions)
            else len(lines)
        )
        block = "\n".join(lines[line_index:next_index]).strip()
        ordinal = len(sections) + 1
        sections.append(
            {
                "section_order": ordinal,
                "heading_level": level,
                "title": title[:500],
                "section_key": section_key(title, ordinal),
                "content_markdown": normalize_newlines(block),
            }
        )

    return sections


def load_id_registry(path: Path) -> dict[str, str]:
    if not path.is_file():
        raise RuntimeError(f"HISTORY_ID_REGISTRY_NOT_FOUND: {path}")

    data = json.loads(path.read_text(encoding="utf-8"))
    entries = data.get("entries")

    if not isinstance(entries, dict):
        raise RuntimeError("HISTORY_ID_REGISTRY_ENTRIES_INVALID")

    mapping: dict[str, str] = {}
    used_ids: set[str] = set()

    for content_hash, entry in entries.items():
        if not re.fullmatch(r"[0-9a-f]{64}", content_hash):
            raise RuntimeError(
                f"HISTORY_ID_REGISTRY_INVALID_HASH: {content_hash}"
            )

        if not isinstance(entry, dict):
            raise RuntimeError(
                f"HISTORY_ID_REGISTRY_INVALID_ENTRY: {content_hash}"
            )

        document_id = entry.get("document_id")

        if (
            not isinstance(document_id, str)
            or not DOCUMENT_ID_RE.fullmatch(document_id)
        ):
            raise RuntimeError(
                f"HISTORY_ID_REGISTRY_INVALID_DOCUMENT_ID: {document_id}"
            )

        if document_id in used_ids:
            raise RuntimeError(
                f"HISTORY_ID_REGISTRY_DUPLICATE_DOCUMENT_ID: {document_id}"
            )

        used_ids.add(document_id)
        mapping[content_hash] = document_id

    return mapping


def build_corpus(
    root: Path,
    id_registry: dict[str, str],
) -> tuple[list[CorpusDocument], list[ExtractedFile]]:
    extracted: list[ExtractedFile] = []

    for path in sorted(root.rglob("*"), key=lambda item: str(item).lower()):
        if not path.is_file():
            continue

        relative_path = str(path.relative_to(root)).replace("\\", "/")

        if relative_path in SENSITIVE_EXCLUDED_RELATIVE_PATHS:
            print(f"EXCLUDED_SENSITIVE  : {relative_path}")
            continue

        if path.suffix.lower() not in ALL_SUPPORTED_EXTENSIONS:
            continue

        extracted.append(extract_file(root, path))

    groups: dict[str, list[ExtractedFile]] = defaultdict(list)
    for item in extracted:
        groups[item.content_sha256].append(item)

    corpus: list[CorpusDocument] = []
    ordered_groups = sorted(
        groups.values(),
        key=lambda group: min(item.relative_path.lower() for item in group),
    )

    missing_registry_entries: list[tuple[str, str]] = []

    for variants in ordered_groups:
        variants = sorted(
            variants,
            key=lambda item: item.relative_path.lower(),
        )
        canonical = variants[0]

        document_id = id_registry.get(canonical.content_sha256)

        if document_id is None:
            missing_registry_entries.append(
                (
                    canonical.content_sha256,
                    canonical.relative_path,
                )
            )
            continue

        corpus.append(
            CorpusDocument(
                document_id=document_id,
                canonical=canonical,
                variants=variants,
                sections=parse_sections(canonical.content_markdown),
            )
        )

    if missing_registry_entries:
        details = "\n".join(
            f"{content_hash} | {relative_path}"
            for content_hash, relative_path
            in missing_registry_entries[:20]
        )

        raise RuntimeError(
            "UNREGISTERED_HISTORY_CONTENT_HASHES: "
            f"{len(missing_registry_entries)}\n{details}"
        )

    corpus.sort(key=lambda item: item.document_id)

    return corpus, extracted


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Importuje volně strukturovaný historický korpus MatchMatrix."
    )
    parser.add_argument(
        "--root",
        default=DEFAULT_ARCHIVE_ROOT,
        help="Kořenová složka historického archivu.",
    )
    parser.add_argument(
        "--id-registry",
        default=str(DEFAULT_ID_REGISTRY),
        help="Registr stabilnich vazeb content SHA-256 -> MM-HIS ID.",
    )
    parser.add_argument(
        "--validate-only",
        action="store_true",
        help="Pouze vytvoří manifest a report; nepřipojuje se k databázi.",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Potvrdí databázovou transakci. Bez přepínače se provede rollback.",
    )
    parser.add_argument(
        "--dsn",
        help="PostgreSQL DSN. Lze nahradit standardními PG* proměnnými prostředí.",
    )
    return parser.parse_args()


def report_directory() -> Path:
    project_root = Path(__file__).resolve().parents[2]
    path = project_root / "reports" / "documentation"
    path.mkdir(parents=True, exist_ok=True)
    return path


def write_json_report(prefix: str, payload: dict[str, Any]) -> tuple[Path, Path]:
    directory = report_directory()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    timestamped = directory / f"{prefix}_{stamp}.json"
    latest = directory / f"{prefix}_latest.json"
    text = json.dumps(payload, ensure_ascii=False, indent=2, default=str)
    timestamped.write_text(text, encoding="utf-8")
    latest.write_text(text, encoding="utf-8")
    return timestamped, latest


def write_manifest_csv(corpus: list[CorpusDocument]) -> Path:
    directory = report_directory()
    path = directory / "history_corpus_manifest_latest.csv"

    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "document_id",
                "title",
                "canonical_relative_path",
                "detected_format",
                "document_date",
                "content_sha256",
                "variant_count",
                "section_count",
                "extraction_status",
                "warnings",
            ],
        )
        writer.writeheader()

        for document in corpus:
            writer.writerow(
                {
                    "document_id": document.document_id,
                    "title": document.canonical.title,
                    "canonical_relative_path": document.canonical.relative_path,
                    "detected_format": document.canonical.detected_format,
                    "document_date": document.canonical.document_date,
                    "content_sha256": document.canonical.content_sha256,
                    "variant_count": len(document.variants),
                    "section_count": len(document.sections),
                    "extraction_status": document.canonical.extraction_status,
                    "warnings": ";".join(document.canonical.extraction_warnings),
                }
            )

    return path


def connection_kwargs() -> dict[str, Any]:
    aliases = {
        "host": ("PGHOST", "DB_HOST", "POSTGRES_HOST"),
        "port": ("PGPORT", "DB_PORT", "POSTGRES_PORT"),
        "dbname": ("PGDATABASE", "DB_NAME", "POSTGRES_DB"),
        "user": ("PGUSER", "DB_USER", "POSTGRES_USER"),
        "password": ("PGPASSWORD", "DB_PASSWORD", "POSTGRES_PASSWORD"),
    }
    values: dict[str, Any] = {}

    for key, names in aliases.items():
        for name in names:
            value = os.environ.get(name)
            if value:
                values[key] = value
                break

    if "port" in values:
        values["port"] = int(values["port"])

    return values


def connect_database(dsn: str | None):
    try:
        import psycopg
    except ImportError as exc:
        raise RuntimeError(
            "Chybí psycopg. Nainstaluj: py -3.14 -m pip install \"psycopg[binary]\""
        ) from exc

    if dsn:
        return psycopg.connect(dsn)

    kwargs = connection_kwargs()
    if not kwargs:
        raise RuntimeError(
            "Chybí databázové připojení. Použij --dsn nebo PG* proměnné prostředí."
        )
    return psycopg.connect(**kwargs)


def jsonb(value: Any):
    from psycopg.types.json import Jsonb
    return Jsonb(value)


def create_import_run(connection, root: Path, files_scanned: int) -> int:
    with connection.cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO documentation.import_runs (
                import_status,
                import_mode,
                source_root,
                files_scanned,
                files_imported,
                files_updated,
                files_skipped,
                files_failed,
                notes,
                details
            )
            VALUES (
                'RUNNING',
                'FULL_SCAN',
                %s,
                %s,
                0,
                0,
                0,
                0,
                %s,
                %s
            )
            RETURNING import_run_pk
            """,
            (
                str(root),
                files_scanned,
                "Import nekanonického historického korpusu MatchMatrix.",
                jsonb({"engine": ENGINE_VERSION}),
            ),
        )
        return int(cursor.fetchone()[0])


def fetch_document(connection, repo_path: str) -> dict[str, Any] | None:
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT
                document_pk,
                document_id,
                current_version_label,
                current_status
            FROM documentation.documents
            WHERE repo_relative_path = %s
            """,
            (repo_path,),
        )
        row = cursor.fetchone()

    if not row:
        return None

    return {
        "document_pk": row[0],
        "document_id": row[1],
        "current_version_label": row[2],
        "current_status": row[3],
    }


def fetch_current_version(connection, document_pk: int) -> dict[str, Any] | None:
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT version_pk, version_label, content_hash_sha256
            FROM documentation.document_versions
            WHERE document_pk = %s
              AND is_current = true
            LIMIT 1
            """,
            (document_pk,),
        )
        row = cursor.fetchone()

    if not row:
        return None

    return {
        "version_pk": row[0],
        "version_label": row[1],
        "content_hash_sha256": row[2],
    }


def next_version_label(current: str | None) -> str:
    if not current:
        return "1.0"
    match = re.fullmatch(r"(\d+)\.(\d+)", current)
    if not match:
        return "1.1"
    return f"{match.group(1)}.{int(match.group(2)) + 1}"


def upsert_corpus_document(
    connection,
    import_run_pk: int,
    archive_root: Path,
    document: CorpusDocument,
) -> str:
    canonical = document.canonical
    repo_path = (
        "archive/09_HISTORY/historie_25062026/"
        + canonical.relative_path
    )
    existing = fetch_document(connection, repo_path)
    metadata = document.metadata
    metadata["archive_root"] = str(archive_root)
    metadata["file_metadata"] = canonical.metadata

    if existing is None:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO documentation.documents (
                    document_id,
                    document_type,
                    title,
                    edition,
                    language_code,
                    current_version_label,
                    current_status,
                    category,
                    canonical_filename,
                    repo_relative_path,
                    source_of_truth,
                    is_active,
                    metadata
                )
                VALUES (
                    %s,
                    'HIS',
                    %s,
                    'HISTORY',
                    'cs',
                    '1.0',
                    'ARCHIVED',
                    'HISTORICAL_CORPUS',
                    %s,
                    %s,
                    'FILE',
                    true,
                    %s
                )
                RETURNING document_pk
                """,
                (
                    document.document_id,
                    canonical.title,
                    canonical.absolute_path.name,
                    repo_path,
                    jsonb(metadata),
                ),
            )
            document_pk = int(cursor.fetchone()[0])
        version_label = "1.0"
        state = "INSERTED"

    else:
        document_pk = int(existing["document_pk"])
        current = fetch_current_version(connection, document_pk)

        if current and current["content_hash_sha256"] == canonical.content_sha256:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    UPDATE documentation.documents
                    SET
                        title = %s,
                        canonical_filename = %s,
                        metadata = %s,
                        updated_at = now()
                    WHERE document_pk = %s
                    """,
                    (
                        canonical.title,
                        canonical.absolute_path.name,
                        jsonb(metadata),
                        document_pk,
                    ),
                )
            return "SKIPPED_SAME_CONTENT"

        version_label = next_version_label(
            current["version_label"] if current else existing["current_version_label"]
        )
        with connection.cursor() as cursor:
            cursor.execute(
                """
                UPDATE documentation.document_versions
                SET is_current = false
                WHERE document_pk = %s
                  AND is_current = true
                """,
                (document_pk,),
            )
            cursor.execute(
                """
                UPDATE documentation.documents
                SET
                    title = %s,
                    current_version_label = %s,
                    current_status = 'ARCHIVED',
                    canonical_filename = %s,
                    metadata = %s,
                    updated_at = now()
                WHERE document_pk = %s
                """,
                (
                    canonical.title,
                    version_label,
                    canonical.absolute_path.name,
                    jsonb(metadata),
                    document_pk,
                ),
            )
        state = "UPDATED"

    version_metadata = {
        **metadata,
        "source_absolute_path": str(canonical.absolute_path),
    }

    with connection.cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO documentation.document_versions (
                document_pk,
                import_run_pk,
                version_label,
                version_status,
                content_markdown,
                content_hash_sha256,
                source_filename,
                source_file_path,
                source_modified_at,
                change_summary,
                is_current,
                imported_by,
                metadata
            )
            VALUES (
                %s,
                %s,
                %s,
                'ARCHIVED',
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                true,
                'A25_HISTORY_CORPUS_IMPORT',
                %s
            )
            RETURNING version_pk
            """,
            (
                document_pk,
                import_run_pk,
                version_label,
                canonical.content_markdown,
                canonical.content_sha256,
                canonical.absolute_path.name,
                str(canonical.absolute_path),
                canonical.modified_at,
                "Import historického korpusu bez obsahové standardizace.",
                jsonb(version_metadata),
            ),
        )
        version_pk = int(cursor.fetchone()[0])

    for section in document.sections:
        section_content = section["content_markdown"]
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO documentation.document_sections (
                    version_pk,
                    section_key,
                    section_order,
                    heading_level,
                    title,
                    content_markdown,
                    content_hash_sha256,
                    metadata
                )
                VALUES (
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s
                )
                """,
                (
                    version_pk,
                    section["section_key"],
                    section["section_order"],
                    section["heading_level"],
                    section["title"],
                    section_content,
                    sha256_bytes(section_content.encode("utf-8")),
                    jsonb({"source": "A25_HISTORY_CORPUS_IMPORT"}),
                ),
            )

    with connection.cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO documentation.document_status_history (
                document_pk,
                version_pk,
                previous_status,
                new_status,
                change_reason,
                changed_by,
                metadata
            )
            VALUES (
                %s,
                %s,
                %s,
                'ARCHIVED',
                %s,
                'A25_HISTORY_CORPUS_IMPORT',
                %s
            )
            """,
            (
                document_pk,
                version_pk,
                None if state == "INSERTED" else "ARCHIVED",
                "HISTORY_CORPUS_IMPORTED",
                jsonb({"engine": ENGINE_VERSION}),
            ),
        )

    return state


def finalize_import_run(
    connection,
    import_run_pk: int,
    *,
    status: str,
    imported: int,
    updated: int,
    skipped: int,
    failed: int,
    details: dict[str, Any],
) -> None:
    with connection.cursor() as cursor:
        cursor.execute(
            """
            UPDATE documentation.import_runs
            SET
                finished_at = now(),
                import_status = %s,
                files_imported = %s,
                files_updated = %s,
                files_skipped = %s,
                files_failed = %s,
                details = %s
            WHERE import_run_pk = %s
            """,
            (
                status,
                imported,
                updated,
                skipped,
                failed,
                jsonb(details),
                import_run_pk,
            ),
        )


def manifest_payload(
    root: Path,
    corpus: list[CorpusDocument],
    extracted: list[ExtractedFile],
) -> dict[str, Any]:
    warnings = [
        {
            "relative_path": item.relative_path,
            "warnings": item.extraction_warnings,
        }
        for item in extracted
        if item.extraction_warnings
    ]

    return {
        "engine": ENGINE_VERSION,
        "generated_at": utc_now().isoformat(),
        "archive_root": str(root),
        "summary": {
            "source_files": len(extracted),
            "corpus_documents": len(corpus),
            "content_duplicate_groups": sum(
                1 for document in corpus if len(document.variants) > 1
            ),
            "source_files_in_duplicate_groups": sum(
                len(document.variants)
                for document in corpus
                if len(document.variants) > 1
            ),
            "sections": sum(len(document.sections) for document in corpus),
            "warnings": len(warnings),
        },
        "documents": [
            {
                "document_id": document.document_id,
                "title": document.canonical.title,
                "canonical_relative_path": document.canonical.relative_path,
                "detected_format": document.canonical.detected_format,
                "document_date": document.canonical.document_date,
                "content_sha256": document.canonical.content_sha256,
                "variant_count": len(document.variants),
                "source_variants": [
                    item.relative_path for item in document.variants
                ],
                "section_count": len(document.sections),
                "extraction_status": document.canonical.extraction_status,
                "warnings": document.canonical.extraction_warnings,
            }
            for document in corpus
        ],
        "warnings": warnings,
    }


def main() -> int:
    args = parse_args()
    root = Path(args.root)
    id_registry_path = Path(args.id_registry)

    print("MATCHMATRIX HISTORY CORPUS DATABASE IMPORT")
    print("=" * 79)
    print(f"ENGINE             : {ENGINE_VERSION}")
    print(f"ARCHIVE ROOT       : {root}")
    print(f"ID REGISTRY        : {id_registry_path}")
    print(
        "MODE               : "
        + ("VALIDATE_ONLY" if args.validate_only else ("APPLY" if args.apply else "DRY_RUN"))
    )
    print()

    if not root.is_dir():
        print(f"BLOCKED: Archivní složka neexistuje: {root}", file=sys.stderr)
        return 2

    id_registry = load_id_registry(id_registry_path)
    corpus, extracted = build_corpus(root, id_registry)
    payload = manifest_payload(root, corpus, extracted)
    report_path, latest_path = write_json_report("history_corpus_manifest", payload)
    csv_path = write_manifest_csv(corpus)

    summary = payload["summary"]
    print("KORPUS")
    print("-" * 79)
    for key, value in summary.items():
        print(f"{key:<36}: {value}")
    print()
    print(f"JSON MANIFEST      : {report_path}")
    print(f"LATEST JSON        : {latest_path}")
    print(f"CSV MANIFEST       : {csv_path}")

    if args.validate_only:
        final = (
            "HISTORY_CORPUS_VALIDATED_WITH_WARNINGS"
            if summary["warnings"]
            else "HISTORY_CORPUS_VALIDATED"
        )
        print(f"FINAL STATUS       : {final}")
        return 0

    connection = connect_database(args.dsn)
    imported = 0
    updated = 0
    skipped = 0
    failed = 0

    try:
        import_run_pk = create_import_run(connection, root, len(extracted))
        print()
        print(f"IMPORT RUN ID      : {import_run_pk}")
        print()

        for index, document in enumerate(corpus, start=1):
            try:
                state = upsert_corpus_document(
                    connection,
                    import_run_pk,
                    root,
                    document,
                )
                if state == "INSERTED":
                    imported += 1
                elif state == "UPDATED":
                    updated += 1
                else:
                    skipped += 1

                print(
                    f"{document.document_id} | {state:<20} | "
                    f"variants={len(document.variants):>2} | "
                    f"sections={len(document.sections):>3} | "
                    f"{document.canonical.relative_path}"
                )
            except Exception:
                failed += 1
                raise

        details = {
            "engine": ENGINE_VERSION,
            "archive_root": str(root),
            "corpus_documents": len(corpus),
            "source_files": len(extracted),
            "warnings": payload["warnings"],
        }
        final_db_status = "DONE_WITH_WARNINGS" if summary["warnings"] else "DONE"
        finalize_import_run(
            connection,
            import_run_pk,
            status=final_db_status,
            imported=imported,
            updated=updated,
            skipped=skipped,
            failed=failed,
            details=details,
        )

        if args.apply:
            connection.commit()
            final_status = "HISTORY_CORPUS_IMPORT_APPLIED"
        else:
            connection.rollback()
            final_status = "HISTORY_CORPUS_IMPORT_DRY_RUN_READY"

    except Exception as exc:
        connection.rollback()
        print(f"FAILED: {exc}", file=sys.stderr)
        return 1
    finally:
        connection.close()

    print()
    print("SOUHRN")
    print("-" * 79)
    print(f"documents_inserted                 : {imported}")
    print(f"documents_updated                  : {updated}")
    print(f"documents_skipped                  : {skipped}")
    print(f"documents_failed                   : {failed}")
    print(f"warnings                           : {summary['warnings']}")
    print(f"FINAL STATUS                       : {final_status}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
